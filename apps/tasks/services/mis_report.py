# D:\CLIENT PROJECT\employee management system bos\employee_management_system\apps\tasks\services\mis_report.py
from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, Q, QuerySet
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.html import strip_tags
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)
User = get_user_model()

IST = ZoneInfo(getattr(settings, "MIS_REPORT_TIME_ZONE", "Asia/Kolkata"))

DEFAULT_PRIMARY_RECIPIENTS = ["pankaj@blueoceansteels.com"]
DEFAULT_CC_RECIPIENTS = ["amreen@blueoceansteels.com"]

DEFAULT_TASK_MODEL_NAMES = ("Checklist", "Delegation", "HelpTicket")

# ---------------------------------------------------------------------
# MIS TEAM GROUPING CONFIG
# ---------------------------------------------------------------------
# Presentation-only grouping.
# These names should match the database display names from user.get_full_name().
# This mapping does NOT affect MIS aggregation, planned count, actual count,
# on-time count, score formula, task query filters, or email data flow.
# ---------------------------------------------------------------------
MIS_TEAM_MAPPING = {
    "MDO Team": [
        "Akshay Barangule",
        "Saurabh Kumavat",
        "Amreen Mulla",
        "Kajal Adagale",
        "Vaishnavi Lavand",
        "Dnyaneshwar Kumavat",
        "Rohit Gujar",
        "Smruti Lokhande",
    ],
    "Purchase & Logistics Team": [
        "Rahul Nevase",
        "Sharyu Patil",
        "Yogesh Shinde",
        "Shrikant Gawalwad",
    ],
    "Marketing Team": [
        "Pratik Khandke",
        "Sushant Nanaware",
        "Manoj Dhole",
        "Dinesh Kokate",
        "Gauri Thorat",
    ],
}

OTHER_ACTIVE_USERS_GROUP_NAME = "Other Active Users / Unassigned Team"

# These users can receive the email but must not appear inside employee MIS table.
DEFAULT_EXCLUDE_USERNAMES = ["admin", "pankaj"]
DEFAULT_EXCLUDE_EMAILS = ["admin@gmail.com", "pankaj@blueoceansteels.com"]
DEFAULT_EXCLUDE_FULL_NAMES = ["Pankaj Jain", "Pankaj Sir"]

COMPLETED_STATUS_BY_MODEL = {
    "Checklist": ("Completed",),
    "Delegation": ("Completed",),
    "HelpTicket": ("Closed",),
}

COMPLETED_DATETIME_FIELDS_BY_MODEL = {
    "Checklist": ("completed_at",),
    "Delegation": ("completed_at",),
    "HelpTicket": ("resolved_at",),
}


@dataclass(frozen=True)
class TaskModelConfig:
    model_name: str
    model: Any
    assignee_field: str
    planned_field: str
    status_field: str
    completed_datetime_field: str
    completed_statuses: Tuple[str, ...]
    skipped_field: Optional[str]


def _safe_console_text(value: object) -> str:
    try:
        text = "" if value is None else str(value)
    except Exception:
        text = repr(value)

    try:
        return text.encode("utf-8", errors="replace").decode("utf-8", errors="replace")
    except Exception:
        return text


def _list_setting(name: str, default: Sequence[str]) -> List[str]:
    raw = getattr(settings, name, default)

    if raw is None:
        return list(default)

    if isinstance(raw, str):
        return [part.strip() for part in raw.split(",") if part.strip()]

    try:
        return [str(part).strip() for part in raw if str(part).strip()]
    except Exception:
        return list(default)


def _dedupe_emails(emails: Sequence[str] | None) -> List[str]:
    seen = set()
    out: List[str] = []

    for email in emails or []:
        value = (email or "").strip()
        if not value or "@" not in value:
            continue

        key = value.lower()
        if key in seen:
            continue

        seen.add(key)
        out.append(value)

    return out


def _from_email() -> str:
    return (
        getattr(settings, "DEFAULT_FROM_EMAIL", None)
        or getattr(settings, "EMAIL_HOST_USER", None)
        or "BOS Lakshya ERP <no-reply@blueoceansteels.com>"
    )


def _field_names(model: Any) -> set[str]:
    return {field.name for field in model._meta.fields}


def _first_existing_field(model: Any, candidates: Sequence[str]) -> Optional[str]:
    names = _field_names(model)

    for field_name in candidates:
        if field_name in names:
            return field_name

    return None


def _get_model_config(model_name: str) -> Optional[TaskModelConfig]:
    try:
        model = apps.get_model("tasks", model_name)
    except LookupError:
        logger.warning(_safe_console_text(f"[MIS] Model not found: tasks.{model_name}"))
        return None

    assignee_field = _first_existing_field(model, ("assign_to", "assigned_to"))
    planned_field = _first_existing_field(model, ("planned_date",))
    status_field = _first_existing_field(model, ("status",))
    completed_datetime_field = _first_existing_field(
        model,
        COMPLETED_DATETIME_FIELDS_BY_MODEL.get(
            model_name,
            ("completed_at", "resolved_at", "closed_at"),
        ),
    )
    skipped_field = _first_existing_field(model, ("is_skipped_due_to_leave",))

    missing_fields = []

    if not assignee_field:
        missing_fields.append("assign_to")
    if not planned_field:
        missing_fields.append("planned_date")
    if not status_field:
        missing_fields.append("status")
    if not completed_datetime_field:
        missing_fields.append("completed_at/resolved_at")

    if missing_fields:
        logger.warning(
            _safe_console_text(
                f"[MIS] Skipping {model_name}. Missing fields: {', '.join(missing_fields)}"
            )
        )
        return None

    return TaskModelConfig(
        model_name=model_name,
        model=model,
        assignee_field=assignee_field,
        planned_field=planned_field,
        status_field=status_field,
        completed_datetime_field=completed_datetime_field,
        completed_statuses=COMPLETED_STATUS_BY_MODEL.get(
            model_name,
            ("Completed", "Closed"),
        ),
        skipped_field=skipped_field,
    )


def get_available_task_configs() -> List[TaskModelConfig]:
    model_names = getattr(settings, "MIS_TASK_MODEL_NAMES", DEFAULT_TASK_MODEL_NAMES)

    configs: List[TaskModelConfig] = []

    for model_name in model_names:
        config = _get_model_config(str(model_name))
        if config:
            configs.append(config)

    return configs


def get_week_bounds_ist(
    *,
    anchor_date: Optional[date] = None,
    week_selector: str = "current",
) -> Tuple[datetime, datetime, date, date]:
    """
    Week rule:
        Monday to Saturday.

    For Monday 10:30 AM scheduled reports, use --week last.
    """
    if anchor_date is None:
        anchor_date = timezone.localtime(timezone.now(), IST).date()

    current_monday = anchor_date - timedelta(days=anchor_date.weekday())

    selector = (week_selector or "current").strip().lower()

    if selector == "current":
        week_start = current_monday
    elif selector == "last":
        week_start = current_monday - timedelta(days=7)
    else:
        raise ValueError("week_selector must be either 'current' or 'last'")

    week_end = week_start + timedelta(days=5)

    start_dt = timezone.make_aware(datetime.combine(week_start, time.min), IST)
    end_dt = timezone.make_aware(datetime.combine(week_end + timedelta(days=1), time.min), IST)

    return start_dt, end_dt, week_start, week_end


def _case_insensitive_q(field_name: str, values: Sequence[str]) -> Q:
    q = Q(pk__in=[])

    for value in values:
        value = (value or "").strip()
        if value:
            q |= Q(**{f"{field_name}__iexact": value})

    return q


def _display_name(user: Any, fallback_user_id: int) -> str:
    if not user:
        return f"User #{fallback_user_id}"

    try:
        full_name = (user.get_full_name() or "").strip()
        if full_name:
            return full_name
    except Exception:
        pass

    username = (getattr(user, "username", "") or "").strip()
    if username:
        return username

    return f"User #{fallback_user_id}"


def _normalize_person_name(value: object) -> str:
    """
    Normalize employee/team names for safe display grouping.

    Example:
        " Akshay Barangule " -> "akshay barangule"
        "AKSHAY BARANGULE"   -> "akshay barangule"

    Presentation-only helper.
    Does not affect MIS aggregation.
    """
    return " ".join(str(value or "").strip().lower().split())


def _group_employees_by_team(
    employees: Sequence[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Group already-computed employee MIS rows into hardcoded teams.

    Production safety:
        - Does not query tasks again.
        - Does not recalculate planned.
        - Does not recalculate actual.
        - Does not recalculate on-time.
        - Does not recalculate scores.
        - Does not remove active employees silently.

    Users not present in MIS_TEAM_MAPPING are returned separately as
    other_active_users / unmapped_employees for presentation.
    """
    employees_by_name: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for row in employees:
        key = _normalize_person_name(row.get("employee_name"))
        if key:
            employees_by_name[key].append(row)

    used_employee_ids = set()
    team_groups: List[Dict[str, Any]] = []

    for team_name, configured_names in MIS_TEAM_MAPPING.items():
        team_rows: List[Dict[str, Any]] = []

        for configured_name in configured_names:
            key = _normalize_person_name(configured_name)

            for row in employees_by_name.get(key, []):
                employee_id = row.get("employee_id")

                if employee_id in used_employee_ids:
                    continue

                team_rows.append(row)
                used_employee_ids.add(employee_id)

        team_groups.append(
            {
                "team_name": team_name,
                "employees": team_rows,
                "employee_count": len(team_rows),
                "has_rows": bool(team_rows),
            }
        )

    other_active_users = [
        row
        for row in employees
        if row.get("employee_id") not in used_employee_ids
    ]

    return team_groups, other_active_users


def _eligible_active_employee_users() -> List[Any]:
    """
    MIS employee eligibility.

    Include:
        - active users only
        - superusers if settings.MIS_EXCLUDE_SUPERUSERS is False
        - staff users if settings.MIS_EXCLUDE_STAFF is False

    Exclude:
        - optional superusers if configured
        - optional staff if configured
        - configured admin/system usernames
        - Pankaj Sir from the MIS employee table
        - configured excluded emails/full names

    Important:
        Pankaj can still receive email. This only excludes him from the report table.
    """
    qs = User.objects.filter(is_active=True)

    if getattr(settings, "MIS_EXCLUDE_SUPERUSERS", True):
        qs = qs.filter(is_superuser=False)

    if getattr(settings, "MIS_EXCLUDE_STAFF", False):
        qs = qs.filter(is_staff=False)

    exclude_usernames = _list_setting("MIS_EXCLUDE_USERNAMES", DEFAULT_EXCLUDE_USERNAMES)
    exclude_emails = _list_setting("MIS_EXCLUDE_EMAILS", DEFAULT_EXCLUDE_EMAILS)
    exclude_full_names = _list_setting("MIS_EXCLUDE_FULL_NAMES", DEFAULT_EXCLUDE_FULL_NAMES)

    if exclude_usernames:
        qs = qs.exclude(_case_insensitive_q("username", exclude_usernames))

    if exclude_emails:
        qs = qs.exclude(_case_insensitive_q("email", exclude_emails))

    users = list(qs.order_by("first_name", "last_name", "username", "id"))

    exclude_full_names_lower = {
        item.strip().lower()
        for item in exclude_full_names
        if item and item.strip()
    }

    final_users = []

    for user in users:
        full_name = (user.get_full_name() or "").strip().lower()
        username = (getattr(user, "username", "") or "").strip().lower()
        email = (getattr(user, "email", "") or "").strip().lower()

        if full_name and full_name in exclude_full_names_lower:
            continue

        if username and username in exclude_full_names_lower:
            continue

        if email and email in exclude_full_names_lower:
            continue

        final_users.append(user)

    return final_users


def _base_queryset(
    config: TaskModelConfig,
    *,
    start_dt: datetime,
    end_dt: datetime,
    eligible_employee_ids: Sequence[int],
) -> QuerySet:
    """
    Base MIS queryset.

    Critical production rule:
        Only eligible active employees are counted.

    This excludes:
        - inactive employees
        - Pankaj Sir
        - configured admin/system accounts
    """
    filters = {
        f"{config.planned_field}__gte": start_dt,
        f"{config.planned_field}__lt": end_dt,
        f"{config.assignee_field}__isnull": False,
        f"{config.assignee_field}_id__in": list(eligible_employee_ids),
    }

    qs = config.model.objects.filter(**filters)

    if config.skipped_field:
        qs = qs.filter(**{config.skipped_field: False})

    return qs


def _inactive_assignee_queryset(
    config: TaskModelConfig,
    *,
    start_dt: datetime,
    end_dt: datetime,
) -> QuerySet:
    filters = {
        f"{config.planned_field}__gte": start_dt,
        f"{config.planned_field}__lt": end_dt,
        f"{config.assignee_field}__isnull": False,
        f"{config.assignee_field}__is_active": False,
    }

    qs = config.model.objects.filter(**filters)

    if config.skipped_field:
        qs = qs.filter(**{config.skipped_field: False})

    return qs


def _excluded_active_assignee_queryset(
    config: TaskModelConfig,
    *,
    start_dt: datetime,
    end_dt: datetime,
    eligible_employee_ids: Sequence[int],
) -> QuerySet:
    """
    Audit-only queryset.

    Counts rows assigned to active users who were excluded from MIS,
    such as Pankaj Sir or configured admin/system users.
    """
    filters = {
        f"{config.planned_field}__gte": start_dt,
        f"{config.planned_field}__lt": end_dt,
        f"{config.assignee_field}__isnull": False,
        f"{config.assignee_field}__is_active": True,
    }

    qs = config.model.objects.filter(**filters).exclude(
        **{f"{config.assignee_field}_id__in": list(eligible_employee_ids)}
    )

    if config.skipped_field:
        qs = qs.filter(**{config.skipped_field: False})

    return qs


def _completed_queryset(base_qs: QuerySet, config: TaskModelConfig) -> QuerySet:
    return base_qs.filter(
        **{
            f"{config.status_field}__in": config.completed_statuses,
            f"{config.completed_datetime_field}__isnull": False,
        }
    )


def _count_by_assignee(qs: QuerySet, assignee_field: str) -> Dict[int, int]:
    user_id_key = f"{assignee_field}_id"

    counts: Dict[int, int] = {}

    for row in qs.values(user_id_key).annotate(total=Count("id")):
        user_id = row.get(user_id_key)
        if user_id:
            counts[int(user_id)] = int(row.get("total") or 0)

    return counts


def _is_on_time(planned_dt: datetime, completed_dt: datetime) -> bool:
    if not planned_dt or not completed_dt:
        return False

    planned_date_ist = timezone.localtime(planned_dt, IST).date()
    completed_date_ist = timezone.localtime(completed_dt, IST).date()

    return completed_date_ist <= planned_date_ist


def _count_on_time_by_assignee(
    completed_qs: QuerySet,
    config: TaskModelConfig,
) -> Dict[int, int]:
    counts: Dict[int, int] = defaultdict(int)

    qs = completed_qs.only(
        "id",
        config.assignee_field,
        config.planned_field,
        config.completed_datetime_field,
    )

    for obj in qs.iterator(chunk_size=1000):
        user_id = getattr(obj, f"{config.assignee_field}_id", None)
        planned_dt = getattr(obj, config.planned_field, None)
        completed_dt = getattr(obj, config.completed_datetime_field, None)

        if not user_id:
            continue

        if _is_on_time(planned_dt, completed_dt):
            counts[int(user_id)] += 1

    return dict(counts)


def _score(
    *,
    actual: int,
    planned: int,
    formula: str,
) -> Decimal:
    if planned <= 0:
        return Decimal("0.00")

    formula = (formula or "variance").strip().lower()

    if formula == "completion_pct":
        value = (Decimal(actual) / Decimal(planned)) * Decimal("100")
    else:
        value = ((Decimal(actual) - Decimal(planned)) / Decimal(planned)) * Decimal("100")

    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _format_score(value: Decimal) -> str:
    value = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    if value == value.to_integral():
        return str(int(value))

    return str(value)


def _week_number(week_start: date) -> int:
    return int(week_start.isocalendar().week)


def _week_year(week_start: date) -> int:
    return int(week_start.isocalendar().year)


def build_mis_report_dataset(
    *,
    anchor_date: Optional[date] = None,
    week_selector: str = "current",
    formula: Optional[str] = None,
) -> Dict[str, Any]:
    formula = formula or getattr(settings, "MIS_SCORE_FORMULA", "variance")

    start_dt, end_dt, week_start, week_end = get_week_bounds_ist(
        anchor_date=anchor_date,
        week_selector=week_selector,
    )

    generated_at = timezone.localtime(timezone.now(), IST)
    report_date = generated_at.date()

    active_employee_count_in_db = User.objects.filter(is_active=True).count()
    inactive_employee_count_in_db = User.objects.filter(is_active=False).count()

    eligible_users = _eligible_active_employee_users()
    eligible_employee_ids = [user.id for user in eligible_users]
    eligible_users_by_id = {user.id: user for user in eligible_users}

    employee_stats: Dict[int, Dict[str, int]] = defaultdict(
        lambda: {
            "planned": 0,
            "actual": 0,
            "on_time_actual": 0,
        }
    )

    # Show every eligible active employee, even if the selected week has 0 tasks.
    for user_id in eligible_employee_ids:
        employee_stats[user_id]

    model_breakdown: Dict[str, Dict[str, int]] = {}

    configs = get_available_task_configs()

    for config in configs:
        base_qs = _base_queryset(
            config,
            start_dt=start_dt,
            end_dt=end_dt,
            eligible_employee_ids=eligible_employee_ids,
        )

        inactive_qs = _inactive_assignee_queryset(
            config,
            start_dt=start_dt,
            end_dt=end_dt,
        )

        excluded_active_qs = _excluded_active_assignee_queryset(
            config,
            start_dt=start_dt,
            end_dt=end_dt,
            eligible_employee_ids=eligible_employee_ids,
        )

        completed_qs = _completed_queryset(base_qs, config)

        planned_counts = _count_by_assignee(base_qs, config.assignee_field)
        actual_counts = _count_by_assignee(completed_qs, config.assignee_field)
        on_time_counts = _count_on_time_by_assignee(completed_qs, config)

        model_breakdown[config.model_name] = {
            "planned": sum(planned_counts.values()),
            "actual": sum(actual_counts.values()),
            "on_time_actual": sum(on_time_counts.values()),
            "inactive_assignee_tasks_skipped": inactive_qs.count(),
            "excluded_active_assignee_tasks_skipped": excluded_active_qs.count(),
        }

        all_user_ids = set(planned_counts) | set(actual_counts) | set(on_time_counts)

        for user_id in all_user_ids:
            if user_id not in eligible_users_by_id:
                continue

            employee_stats[user_id]["planned"] += planned_counts.get(user_id, 0)
            employee_stats[user_id]["actual"] += actual_counts.get(user_id, 0)
            employee_stats[user_id]["on_time_actual"] += on_time_counts.get(user_id, 0)

    employees: List[Dict[str, Any]] = []

    for user_id, stats in employee_stats.items():
        user = eligible_users_by_id.get(user_id)

        if not user:
            continue

        planned = int(stats["planned"])
        actual = int(stats["actual"])
        on_time_actual = int(stats["on_time_actual"])

        current_week_score = _score(
            actual=actual,
            planned=planned,
            formula=formula,
        )

        on_time_planned = actual

        on_time_score = _score(
            actual=on_time_actual,
            planned=on_time_planned,
            formula=formula,
        )

        employees.append(
            {
                "employee_id": user_id,
                "employee_name": _display_name(user, user_id),
                "planned": planned,
                "actual": actual,
                "current_week_score": _format_score(current_week_score),
                "on_time_planned": on_time_planned,
                "on_time_actual": on_time_actual,
                "on_time_score": _format_score(on_time_score),
            }
        )

    employees.sort(key=lambda item: item["employee_name"].lower())

    # Presentation-only grouping. Existing employee rows and totals remain unchanged.
    team_groups, other_active_users = _group_employees_by_team(employees)

    total_planned = sum(row["planned"] for row in employees)
    total_actual = sum(row["actual"] for row in employees)
    total_on_time_actual = sum(row["on_time_actual"] for row in employees)

    total_score = _score(
        actual=total_actual,
        planned=total_planned,
        formula=formula,
    )

    total_on_time_score = _score(
        actual=total_on_time_actual,
        planned=total_actual,
        formula=formula,
    )

    inactive_assignee_tasks_skipped_total = sum(
        item.get("inactive_assignee_tasks_skipped", 0)
        for item in model_breakdown.values()
    )

    excluded_active_assignee_tasks_skipped_total = sum(
        item.get("excluded_active_assignee_tasks_skipped", 0)
        for item in model_breakdown.values()
    )

    return {
        "title": "MDO",
        "report_date": report_date,
        "report_date_display": report_date.strftime("%d-%m-%Y"),
        "week_number": _week_number(week_start),
        "week_year": _week_year(week_start),
        "week_start": week_start,
        "week_end": week_end,
        "week_start_display": week_start.strftime("%d-%m-%Y"),
        "week_end_display": week_end.strftime("%d-%m-%Y"),
        "generated_at": generated_at,
        "generated_at_display": generated_at.strftime("%d-%m-%Y %I:%M %p"),
        "formula": formula,

        # Existing flat list preserved for backward compatibility.
        "employees": employees,

        # Team-wise presentation grouping.
        "team_groups": team_groups,

        # Preferred presentation name.
        "other_active_users": other_active_users,
        "has_other_active_users": bool(other_active_users),
        "other_active_users_group_name": OTHER_ACTIVE_USERS_GROUP_NAME,

        # Backward-compatible aliases. Kept so older templates/debug scripts do not break.
        "unmapped_employees": other_active_users,
        "has_unmapped_employees": bool(other_active_users),

        "employee_count": len(employees),
        "active_employee_count_in_db": active_employee_count_in_db,
        "inactive_employee_count_in_db": inactive_employee_count_in_db,
        "eligible_active_employee_count": len(eligible_employee_ids),
        "inactive_assignee_tasks_skipped_total": inactive_assignee_tasks_skipped_total,
        "excluded_active_assignee_tasks_skipped_total": excluded_active_assignee_tasks_skipped_total,
        "totals": {
            "planned": total_planned,
            "actual": total_actual,
            "current_week_score": _format_score(total_score),
            "on_time_planned": total_actual,
            "on_time_actual": total_on_time_actual,
            "on_time_score": _format_score(total_on_time_score),
        },
        "model_breakdown": model_breakdown,
        "has_rows": bool(employees),
    }


def build_mis_report_excel(report: Dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for MIS Excel attachment. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "MDO"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    grey_fill = PatternFill("solid", fgColor="BFBFBF")
    yellow_fill = PatternFill("solid", fgColor="F7DF9B")
    teal_fill = PatternFill("solid", fgColor="0F5B66")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    green_fill = PatternFill("solid", fgColor="D9EAD3")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:D1")
    ws["A1"] = "MDO"
    ws["A1"].fill = grey_fill
    ws["A1"].font = Font(bold=True, color="000000", size=14)
    ws["A1"].alignment = center
    ws["A1"].border = border

    headers = ["Doer Name", "Planned", "Actual", "Current\nWeek"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = yellow_fill
        cell.font = Font(bold=True, color="000000", size=12)
        cell.alignment = center
        cell.border = border

    row_no = 3

    if report.get("employees"):
        groups = report.get("team_groups") or [
            {
                "team_name": "Employees",
                "employees": report.get("employees", []),
            }
        ]

        for group in groups:
            group_rows = group.get("employees") or []

            if not group_rows:
                continue

            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=4)
            team_cell = ws.cell(row=row_no, column=1)
            team_cell.value = group.get("team_name") or "Team"
            team_cell.fill = grey_fill
            team_cell.font = Font(bold=True, color="000000", size=12)
            team_cell.alignment = left
            team_cell.border = border

            for col in range(1, 5):
                ws.cell(row=row_no, column=col).border = border

            row_no += 1

            for row in group_rows:
                ws.cell(row=row_no, column=1).value = row["employee_name"]
                ws.cell(row=row_no, column=2).value = row["planned"]
                ws.cell(row=row_no, column=3).value = row["actual"]
                ws.cell(row=row_no, column=4).value = row["current_week_score"]

                ws.cell(row=row_no, column=1).fill = teal_fill
                ws.cell(row=row_no, column=1).font = Font(bold=True, color="FFFFFF", size=11)
                ws.cell(row=row_no, column=1).alignment = left

                for col in range(2, 5):
                    ws.cell(row=row_no, column=col).fill = white_fill
                    ws.cell(row=row_no, column=col).font = Font(color="000000", size=11)
                    ws.cell(row=row_no, column=col).alignment = right

                for col in range(1, 5):
                    ws.cell(row=row_no, column=col).border = border

                row_no += 1

                ws.cell(row=row_no, column=1).value = "work done on time ----->"
                ws.cell(row=row_no, column=2).value = row["on_time_planned"]
                ws.cell(row=row_no, column=3).value = row["on_time_actual"]
                ws.cell(row=row_no, column=4).value = row["on_time_score"]

                for col in range(1, 5):
                    cell = ws.cell(row=row_no, column=col)
                    cell.fill = white_fill
                    cell.font = Font(color="000000", size=11)
                    cell.border = border
                    cell.alignment = left if col == 1 else right

                row_no += 1

                ws.row_dimensions[row_no].height = 8
                row_no += 1

        other_active_users = report.get("other_active_users") or report.get("unmapped_employees") or []

        if other_active_users:
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=4)
            other_cell = ws.cell(row=row_no, column=1)
            other_cell.value = report.get("other_active_users_group_name") or OTHER_ACTIVE_USERS_GROUP_NAME
            other_cell.fill = grey_fill
            other_cell.font = Font(bold=True, color="000000", size=12)
            other_cell.alignment = left
            other_cell.border = border

            for col in range(1, 5):
                ws.cell(row=row_no, column=col).border = border

            row_no += 1

            for row in other_active_users:
                ws.cell(row=row_no, column=1).value = row["employee_name"]
                ws.cell(row=row_no, column=2).value = row["planned"]
                ws.cell(row=row_no, column=3).value = row["actual"]
                ws.cell(row=row_no, column=4).value = row["current_week_score"]

                ws.cell(row=row_no, column=1).fill = teal_fill
                ws.cell(row=row_no, column=1).font = Font(bold=True, color="FFFFFF", size=11)
                ws.cell(row=row_no, column=1).alignment = left

                for col in range(2, 5):
                    ws.cell(row=row_no, column=col).fill = white_fill
                    ws.cell(row=row_no, column=col).font = Font(color="000000", size=11)
                    ws.cell(row=row_no, column=col).alignment = right

                for col in range(1, 5):
                    ws.cell(row=row_no, column=col).border = border

                row_no += 1

                ws.cell(row=row_no, column=1).value = "work done on time ----->"
                ws.cell(row=row_no, column=2).value = row["on_time_planned"]
                ws.cell(row=row_no, column=3).value = row["on_time_actual"]
                ws.cell(row=row_no, column=4).value = row["on_time_score"]

                for col in range(1, 5):
                    cell = ws.cell(row=row_no, column=col)
                    cell.fill = white_fill
                    cell.font = Font(color="000000", size=11)
                    cell.border = border
                    cell.alignment = left if col == 1 else right

                row_no += 1

                ws.row_dimensions[row_no].height = 8
                row_no += 1

        ws.cell(row=row_no, column=1).value = "Total"
        ws.cell(row=row_no, column=2).value = report["totals"]["planned"]
        ws.cell(row=row_no, column=3).value = report["totals"]["actual"]
        ws.cell(row=row_no, column=4).value = report["totals"]["current_week_score"]

        for col in range(1, 5):
            cell = ws.cell(row=row_no, column=col)
            cell.fill = green_fill
            cell.font = Font(bold=True, color="000000", size=11)
            cell.border = border
            cell.alignment = left if col == 1 else right

        row_no += 1

        ws.cell(row=row_no, column=1).value = "work done on time ----->"
        ws.cell(row=row_no, column=2).value = report["totals"]["on_time_planned"]
        ws.cell(row=row_no, column=3).value = report["totals"]["on_time_actual"]
        ws.cell(row=row_no, column=4).value = report["totals"]["on_time_score"]

        for col in range(1, 5):
            cell = ws.cell(row=row_no, column=col)
            cell.fill = white_fill
            cell.font = Font(color="000000", size=11)
            cell.border = border
            cell.alignment = left if col == 1 else right

        row_no += 2
    else:
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=4)
        cell = ws.cell(row=row_no, column=1)
        cell.value = "No active eligible employees found for MIS report."
        cell.fill = white_fill
        cell.alignment = center
        cell.border = border
        cell.font = Font(size=11)
        row_no += 2

    ws.cell(row=row_no, column=1).value = f"Report Date: {report['report_date_display']}"
    ws.cell(row=row_no + 1, column=1).value = f"Week Number: {report['week_number']} / {report['week_year']}"
    ws.cell(row=row_no + 2, column=1).value = (
        f"Report Week: {report['week_start_display']} to {report['week_end_display']}"
    )
    ws.cell(row=row_no + 3, column=1).value = f"Generated At: {report['generated_at_display']}"
    ws.cell(row=row_no + 4, column=1).value = (
        "Only active eligible employees are included. "
        "Pankaj Sir, inactive users, and configured admin/system users are excluded."
    )
    ws.cell(row=row_no + 5, column=1).value = (
        "This is a system-generated report from BOS Lakshya ERP."
    )

    for r in range(row_no, row_no + 6):
        ws.cell(row=r, column=1).font = Font(size=11)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 15

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def send_mis_report_email(
    *,
    anchor_date: Optional[date] = None,
    week_selector: str = "current",
    formula: Optional[str] = None,
    to: Optional[Sequence[str]] = None,
    cc: Optional[Sequence[str]] = None,
    dry_run: bool = False,
) -> Dict[str, Any]:
    report = build_mis_report_dataset(
        anchor_date=anchor_date,
        week_selector=week_selector,
        formula=formula,
    )

    subject = (
        "MIS Report - Employee Task Performance "
        f"(Week {report['week_number']}, {report['week_start_display']} to {report['week_end_display']})"
    )

    to_list = _dedupe_emails(
        to or getattr(settings, "MIS_REPORT_TO", DEFAULT_PRIMARY_RECIPIENTS)
    )
    cc_list = _dedupe_emails(
        cc or getattr(settings, "MIS_REPORT_CC", DEFAULT_CC_RECIPIENTS)
    )

    excel_filename = (
        "MIS_Report_MDO_"
        f"Week_{report['week_number']}_"
        f"{report['week_start'].strftime('%Y-%m-%d')}_to_"
        f"{report['week_end'].strftime('%Y-%m-%d')}.xlsx"
    )

    if dry_run:
        return {
            "sent": False,
            "dry_run": True,
            "subject": subject,
            "to": to_list,
            "cc": cc_list,
            "employee_count": report["employee_count"],
            "active_employee_count_in_db": report["active_employee_count_in_db"],
            "inactive_employee_count_in_db": report["inactive_employee_count_in_db"],
            "eligible_active_employee_count": report["eligible_active_employee_count"],
            "inactive_assignee_tasks_skipped_total": report["inactive_assignee_tasks_skipped_total"],
            "excluded_active_assignee_tasks_skipped_total": report["excluded_active_assignee_tasks_skipped_total"],
            "totals": report["totals"],
            "model_breakdown": report["model_breakdown"],
            "excel_attachment": excel_filename,
            "report": report,
        }

    if not to_list:
        raise ValueError("MIS report email requires at least one primary recipient.")

    html_body = render_to_string("email/mis_report.html", report)
    text_body = strip_tags(html_body)

    excel_bytes = build_mis_report_excel(report)

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=_from_email(),
        to=to_list,
        cc=cc_list,
    )
    email.attach_alternative(html_body, "text/html")
    email.attach(
        excel_filename,
        excel_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    email.send(fail_silently=False)

    logger.info(
        _safe_console_text(
            "[MIS] Report email sent. "
            f"week={report['week_start']}..{report['week_end']} "
            f"week_number={report['week_number']} "
            f"employees_in_report={report['employee_count']} "
            f"active_users_in_db={report['active_employee_count_in_db']} "
            f"inactive_users_in_db={report['inactive_employee_count_in_db']} "
            f"inactive_assignee_tasks_skipped={report['inactive_assignee_tasks_skipped_total']} "
            f"excluded_active_assignee_tasks_skipped={report['excluded_active_assignee_tasks_skipped_total']} "
            f"planned={report['totals']['planned']} "
            f"actual={report['totals']['actual']} "
            f"on_time={report['totals']['on_time_actual']} "
            f"attachment={excel_filename} "
            f"to={to_list} cc={cc_list}"
        )
    )

    return {
        "sent": True,
        "dry_run": False,
        "subject": subject,
        "to": to_list,
        "cc": cc_list,
        "employee_count": report["employee_count"],
        "active_employee_count_in_db": report["active_employee_count_in_db"],
        "inactive_employee_count_in_db": report["inactive_employee_count_in_db"],
        "eligible_active_employee_count": report["eligible_active_employee_count"],
        "inactive_assignee_tasks_skipped_total": report["inactive_assignee_tasks_skipped_total"],
        "excluded_active_assignee_tasks_skipped_total": report["excluded_active_assignee_tasks_skipped_total"],
        "week_number": report["week_number"],
        "week_year": report["week_year"],
        "week_start": str(report["week_start"]),
        "week_end": str(report["week_end"]),
        "totals": report["totals"],
        "model_breakdown": report["model_breakdown"],
        "excel_attachment": excel_filename,
    }